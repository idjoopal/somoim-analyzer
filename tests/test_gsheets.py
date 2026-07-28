"""구글 시트 연동 테스트 — 네트워크 무관 (Drive API는 전부 모킹).

가장 중요한 것은 **타입 정규화**다. 구글 변환(xlsx → Sheets → xlsx)을 거치면
날짜 셀이 문자열이나 엑셀 serial로 바뀔 수 있는데, 그대로 흘리면 다운스트림의
`p["posted_at"].month`가 런타임에 터진다. 실제 API 없이도 그 상황을 워크북에
직접 만들어 재현할 수 있다.
"""

from datetime import date, datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from core.excel_builder import (
    _coerce_dt,
    _coerce_iso_date,
    build_excel,
    load_excel_bundle,
)
from core.gsheets import (
    MIME_SHEET,
    MIME_XLSX,
    GSheetsError,
    GoogleSheetsStore,
    default_title,
    parse_credentials,
    parse_sheet_id,
    sheet_url,
)


# ═══════════════════════════════════════════════════════════════
# 순수 함수
# ═══════════════════════════════════════════════════════════════

SHEET_ID = "1AbCdEfGhIjKlMnOpQrStUvWxYz0123456789"


@pytest.mark.parametrize("raw", [
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit#gid=0",
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit",
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}",
    f"https://drive.google.com/file/d/{SHEET_ID}/view?usp=sharing",
    f"https://drive.google.com/open?id={SHEET_ID}",
    SHEET_ID,
    f"  {SHEET_ID}  ",
])
def test_parse_sheet_id_accepts_common_forms(raw):
    assert parse_sheet_id(raw) == SHEET_ID


@pytest.mark.parametrize("bad", ["", "   ", "그냥 텍스트", "https://example.com/", "short"])
def test_parse_sheet_id_rejects_garbage(bad):
    with pytest.raises(GSheetsError):
        parse_sheet_id(bad)


def test_sheet_url_round_trips():
    assert parse_sheet_id(sheet_url(SHEET_ID)) == SHEET_ID


def test_default_title():
    assert default_title("202509-202603") == "다감노_202509-202603_분석"


def test_parse_credentials_accepts_json_string_and_dict():
    info = {"client_email": "a@b.iam.gserviceaccount.com", "private_key": "-----KEY-----"}
    assert parse_credentials(info) == info
    import json
    assert parse_credentials(json.dumps(info)) == info


def test_parse_credentials_reports_missing_fields():
    with pytest.raises(GSheetsError, match="private_key"):
        parse_credentials({"client_email": "a@b.com"})
    with pytest.raises(GSheetsError):
        parse_credentials("not json at all")


# ═══════════════════════════════════════════════════════════════
# 타입 정규화 — 구글 변환이 타입을 바꿔도 버텨야 한다
# ═══════════════════════════════════════════════════════════════

@pytest.mark.parametrize("raw,expected", [
    (datetime(2026, 3, 7, 21, 30), datetime(2026, 3, 7, 21, 30)),
    (date(2026, 3, 7), datetime(2026, 3, 7)),
    ("2026-03-07T21:30:00", datetime(2026, 3, 7, 21, 30)),
    ("2026-03-07 21:30:00", datetime(2026, 3, 7, 21, 30)),
    ("2026-03-07", datetime(2026, 3, 7)),
    ("2026/03/07", datetime(2026, 3, 7)),
    (46088, datetime(2026, 3, 7)),          # 엑셀 serial
])
def test_coerce_dt_handles_every_shape(raw, expected):
    assert _coerce_dt(raw) == expected


@pytest.mark.parametrize("empty", [None, "", "   ", "아무말"])
def test_coerce_dt_returns_none_for_unusable(empty):
    assert _coerce_dt(empty) is None


@pytest.mark.parametrize("raw", [
    date(2026, 3, 7), datetime(2026, 3, 7, 21, 30), "2026-03-07",
    "2026-03-07T00:00:00", 46088,
])
def test_coerce_iso_date_always_yields_iso_string(raw):
    assert _coerce_iso_date(raw) == "2026-03-07"


def test_coerce_iso_date_none_passthrough():
    assert _coerce_iso_date(None) is None
    assert _coerce_iso_date("") is None


# ═══════════════════════════════════════════════════════════════
# 시트 왕복 시뮬레이션 (실제 API 없이)
# ═══════════════════════════════════════════════════════════════

def _sample_posts():
    return [{
        "id": "p1", "author": "닉", "wid": "w1", "title": "[풍경] 출사",
        "body": "본문", "outing_date": "2026-03-07",
        "posted_at": datetime(2026, 3, 1, 12, 0),
        "cat": "A", "cat_label": "공지", "category": "풍경",
        "is_outing": True, "is_canceled": False,
        "likes": 1, "comments": 0, "images": 0,
        "needs_review": False, "review_reason": "",
    }]


def _mangle_types(blob: bytes) -> bytes:
    """구글 변환이 날짜 타입을 바꿔 놓은 상황을 재현한다.

    posted_at → 문자열, outing_date → datetime (Sheets가 날짜로 자동 인식한 경우).
    """
    wb = load_workbook(BytesIO(blob))
    ws = wb["_원본_게시글"]
    header = [c.value for c in ws[1]]
    i_posted = header.index("posted_at") + 1
    i_outing = header.index("outing_date") + 1
    for r in range(2, ws.max_row + 1):
        pv = ws.cell(r, i_posted).value
        if isinstance(pv, datetime):
            ws.cell(r, i_posted).value = pv.isoformat()
        ov = ws.cell(r, i_outing).value
        if isinstance(ov, str) and ov:
            ws.cell(r, i_outing).value = datetime.fromisoformat(ov)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_round_trip_survives_google_type_conversion():
    """날짜 타입이 뒤바뀐 워크북도 정상 복원돼야 한다 — 시트 왕복의 핵심 회귀."""
    blob = build_excel(_sample_posts(), [], 202601, 202612)
    loaded = load_excel_bundle(_mangle_types(blob))

    p = loaded["posts"][0]
    assert isinstance(p["posted_at"], datetime)
    assert p["posted_at"] == datetime(2026, 3, 1, 12, 0)
    assert p["outing_date"] == "2026-03-07"      # 문자열로 되돌아와야 함
    assert p["posted_at"].month == 3             # 예전엔 여기서 AttributeError


def test_normal_excel_round_trip_still_exact():
    """정규화가 기존 엑셀 경로를 망가뜨리지 않는지."""
    blob = build_excel(_sample_posts(), [], 202601, 202612)
    p = load_excel_bundle(blob)["posts"][0]
    assert p["posted_at"] == datetime(2026, 3, 1, 12, 0)
    assert p["outing_date"] == "2026-03-07"


# ═══════════════════════════════════════════════════════════════
# Drive 호출 (모킹)
# ═══════════════════════════════════════════════════════════════

class FakeRequest:
    def __init__(self, result, error=None):
        self._result, self._error = result, error

    def execute(self):
        if self._error:
            raise self._error
        return self._result


class FakeFiles:
    def __init__(self, parent):
        self.parent = parent

    def create(self, **kw):
        self.parent.create_kwargs = kw
        return FakeRequest({"id": SHEET_ID}, self.parent.create_error)

    def export(self, **kw):
        self.parent.export_kwargs = kw
        return FakeRequest(b"xlsx-bytes", self.parent.export_error)


class FakePermissions:
    def __init__(self, parent):
        self.parent = parent

    def create(self, **kw):
        self.parent.perm_kwargs = kw
        return FakeRequest({"id": "perm1"}, self.parent.perm_error)


class FakeDrive:
    """Drive v3 서비스의 최소 대역 — 호출 인자를 기록해 검증한다."""

    def __init__(self, create_error=None, export_error=None, perm_error=None):
        self.create_error, self.export_error, self.perm_error = (
            create_error, export_error, perm_error)
        self.create_kwargs = self.export_kwargs = self.perm_kwargs = None

    def files(self):
        return FakeFiles(self)

    def permissions(self):
        return FakePermissions(self)


class FakeHttpError(Exception):
    """googleapiclient.errors.HttpError 흉내 — `.resp.status`만 있으면 된다."""

    def __init__(self, status):
        super().__init__(f"HTTP {status}")
        self.resp = type("R", (), {"status": status})()


def _store(drive, folder_id=None):
    return GoogleSheetsStore({}, folder_id=folder_id, service=drive)


def test_upload_requests_sheet_conversion():
    drive = FakeDrive()
    file_id, url = _store(drive).upload(b"xlsx", "제목")

    assert file_id == SHEET_ID
    assert url == sheet_url(SHEET_ID)
    body = drive.create_kwargs["body"]
    assert body["mimeType"] == MIME_SHEET     # 변환을 요청해야 시트가 된다
    assert body["name"] == "제목"
    assert "parents" not in body


def test_upload_places_file_in_configured_folder():
    drive = FakeDrive()
    _store(drive, folder_id="FOLDER123").upload(b"xlsx", "제목")
    assert drive.create_kwargs["body"]["parents"] == ["FOLDER123"]


def test_download_exports_as_xlsx():
    drive = FakeDrive()
    data = _store(drive).download(sheet_url(SHEET_ID))

    assert data == b"xlsx-bytes"
    assert drive.export_kwargs["fileId"] == SHEET_ID
    assert drive.export_kwargs["mimeType"] == MIME_XLSX


def test_download_accepts_bare_id():
    drive = FakeDrive()
    _store(drive).download(SHEET_ID)
    assert drive.export_kwargs["fileId"] == SHEET_ID


def test_share_anyone_reader():
    drive = FakeDrive()
    _store(drive).share_anyone_reader(SHEET_ID)
    assert drive.perm_kwargs["body"] == {"type": "anyone", "role": "reader"}


@pytest.mark.parametrize("status,hint", [(401, "권한"), (403, "권한"),
                                          (404, "찾을 수 없"), (429, "너무 잦")])
def test_api_errors_become_actionable_messages(status, hint):
    drive = FakeDrive(create_error=FakeHttpError(status))
    with pytest.raises(GSheetsError, match=hint):
        _store(drive).upload(b"xlsx", "제목")


def test_download_error_is_wrapped():
    drive = FakeDrive(export_error=FakeHttpError(404))
    with pytest.raises(GSheetsError):
        _store(drive).download(SHEET_ID)


def test_upload_then_download_feeds_load_excel_bundle():
    """업로드한 bytes를 그대로 돌려받으면 기존 복원 경로에 바로 들어간다."""
    blob = build_excel(_sample_posts(), [], 202601, 202612)

    class EchoDrive(FakeDrive):
        def files(self):
            outer = self

            class F(FakeFiles):
                def export(self, **kw):
                    outer.export_kwargs = kw
                    return FakeRequest(blob)
            return F(self)

    loaded = load_excel_bundle(_store(EchoDrive()).download(SHEET_ID))
    assert loaded["start_ym"] == 202601 and loaded["end_ym"] == 202612
    assert loaded["posts"][0]["outing_date"] == "2026-03-07"
