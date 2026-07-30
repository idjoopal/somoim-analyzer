"""후기 본문 기반 참석자 추적 — 단위 테스트 (pytest).

실행: 레포 루트에서 `python -m pytest tests/ -q`
"""
import sys
sys.path.insert(0, "/home/user/somoim-analyzer")

from datetime import datetime, date

from core.collector import (
    parse_member_csv,
    build_member_master,
    build_member_candidates,
    extract_attendees,
    parse_review_outing_date,
    annotate_review_attendees,
    match_outings_with_reviews,
)


# ── extract_attendees ──────────────────────────────────────────
def test_extract_basic():
    master = {"정원석", "이하얀", "김민수", "권두흥"}
    body = "정원석 이하얀 김민수 권두흥 모르는분 후기입니다"
    assert extract_attendees(body, "후기", master) == ["정원석", "이하얀", "김민수", "권두흥"]


def test_extract_blacklist_wins_over_master():
    master = {"엄태진", "후기"}  # "후기"가 마스터에 있어도
    assert "후기" not in extract_attendees("엄태진 후기", "", master)


def test_extract_dedup_and_order():
    master = {"엄태진", "이민영"}
    assert extract_attendees("엄태진 이민영 엄태진", "", master) == ["엄태진", "이민영"]


def test_extract_strips_title():
    master = {"정원석", "장미"}
    title = "5월 정출 장미 후기"
    assert extract_attendees(title + " 정원석", title, master) == ["정원석"]


def test_extract_empty_body():
    assert extract_attendees("", "제목", {"김철수"}) == []


# ── parse_member_csv ───────────────────────────────────────────
def test_parse_member_csv_real_and_nick():
    names, n2r, r2n = parse_member_csv("실명,닉네임\n정원석,원석사진\n이하얀,하얀")
    assert {"정원석", "원석사진", "이하얀", "하얀"} <= names
    assert n2r["원석사진"] == "정원석"
    assert r2n["이하얀"] == "하얀"


def test_parse_member_csv_single_col_and_alias():
    names, n2r, _ = parse_member_csv("김민수\n권두흥,두흥,두흥이;권두")
    assert "김민수" in names
    assert n2r["두흥"] == "권두흥"
    assert n2r["두흥이"] == "권두흥" and n2r["권두"] == "권두흥"


# ── build_member_master ────────────────────────────────────────
def test_build_master_frequency_threshold():
    body = "철수 영희 민수"
    posts = [{"cat": "E", "title": "", "body": body, "author": "a"} for _ in range(3)]
    master = build_member_master(posts, [], min_freq=3)
    assert {"철수", "영희", "민수"} <= master


def test_build_master_below_freq_excluded():
    posts = [{"cat": "E", "title": "", "body": "철수 영희", "author": "a"}]
    assert "철수" not in build_member_master(posts, [], min_freq=3)


def test_build_master_extra_names_and_blacklist():
    master = build_member_master([], [], extra_names={"정원석", "후기"})
    assert "정원석" in master
    assert "후기" not in master  # 블랙리스트 제거


# ── parse_review_outing_date (과거 해석) ───────────────────────
def test_review_date_most_recent_past_same_year():
    assert parse_review_outing_date("06.06 식물원 후기", "", datetime(2026, 6, 20)) == date(2026, 6, 6)


def test_review_date_wraps_to_previous_year():
    assert parse_review_outing_date("12.28 송년 후기", "", datetime(2026, 1, 5)) == date(2025, 12, 28)


def test_review_date_explicit_year_trusted():
    assert parse_review_outing_date("2026.06.06 후기", "", datetime(2026, 6, 20)) == date(2026, 6, 6)


def test_review_date_reads_months_late_review():
    # 늦게 쓴 후기의 제목 날짜를 버리면 작성일로 떨어지는데, 그러면 **다섯 달
    # 뒤에 열린 남의 출사**에 붙는다. 실제로 그렇게 붙은 후기가 있었다.
    assert parse_review_outing_date(
        "02.22(일) [인물] 압화 프로필 후기", "", datetime(2026, 7, 30)) == date(2026, 2, 22)


def test_review_date_always_takes_nearest_past():
    # 몇 달이 지났어도 제목의 날짜를 읽는다. 틀린 날짜라면 매칭 창(±7일)이
    # 걸러 내지, 여기서 버려서 작성일로 떨어뜨릴 일이 아니다.
    assert parse_review_outing_date(
        "01.02 후기", "", datetime(2026, 12, 1)) == date(2026, 1, 2)


def test_review_date_compact_yymmdd():
    # `260308[풍경]광양매화마을` — 구분자 없이 붙여 쓴 날짜.
    assert parse_review_outing_date(
        "260308[풍경]광양매화마을", "", datetime(2026, 3, 22)) == date(2026, 3, 8)


def test_review_date_compact_ignores_other_six_digits():
    assert parse_review_outing_date("999999 후기", "", datetime(2026, 6, 20)) is None


def test_review_date_none_when_absent():
    assert parse_review_outing_date("그냥 후기", "", datetime(2026, 6, 20)) is None


# ── annotate_review_attendees (자가검증) ───────────────────────
def test_annotate_flags_when_author_missing():
    posts = [{
        "cat": "E", "title": "후기", "body": "이하얀 와주셨어요",
        "author": "딴사람닉", "posted_at": datetime(2026, 6, 20),
    }]
    annotate_review_attendees(posts, {"정원석", "이하얀"}, {"원석닉": "정원석"})
    assert posts[0]["attendees"] == ["이하얀"]
    assert posts[0]["attendees_needs_review"] is True
    assert "명단에 없음" in posts[0]["attendees_review_reason"]


def test_annotate_ok_when_author_present():
    posts = [{
        "cat": "E", "title": "후기", "body": "정원석 이하얀 함께",
        "author": "원석닉", "posted_at": datetime(2026, 6, 20),
    }]
    annotate_review_attendees(posts, {"정원석", "이하얀"}, {"원석닉": "정원석"})
    assert posts[0]["attendees_needs_review"] is False
    assert posts[0]["attendees"] == ["정원석", "이하얀"]


def test_annotate_flags_empty():
    posts = [{
        "cat": "E", "title": "후기", "body": "다 같이 즐거웠어요",
        "author": "닉", "posted_at": datetime(2026, 6, 20),
    }]
    annotate_review_attendees(posts, {"정원석"}, {})
    assert posts[0]["attendees"] == []
    assert posts[0]["attendees_needs_review"] is True


def test_annotate_canonicalizes_nick_to_real():
    posts = [{
        "cat": "E", "title": "", "body": "원석닉 참석",
        "author": "원석닉", "posted_at": datetime(2026, 6, 20),
    }]
    annotate_review_attendees(posts, {"정원석", "원석닉"}, {"원석닉": "정원석"})
    assert posts[0]["attendees"] == ["정원석"]


# ── match_outings_with_reviews ─────────────────────────────────
def _notice(nid, d, cat, author="공지자", title="", canceled=False):
    return {"id": nid, "cat": "A", "outing_date": d, "category": cat,
            "author": author, "title": title, "is_canceled": canceled,
            "posted_at": datetime.fromisoformat(d + "T00:00:00")}


def _review(rid, attendees, posted, cat=None, author="작성자", rod=None,
            title="", body=""):
    return {"id": rid, "cat": "E", "attendees": list(attendees), "posted_at": posted,
            "category": cat, "review_outing_date": rod, "title": title, "body": body}


def test_match_category_breaks_date_tie():
    notices = [_notice("n1", "2026-06-06", "풍경"), _notice("n2", "2026-06-06", "인물")]
    rev = _review("r1", ["정원석"], datetime(2026, 6, 7), cat="인물", rod="2026-06-06")
    match_outings_with_reviews(notices + [rev])
    by_id = {n["id"]: n for n in notices}
    assert by_id["n2"]["matched_review_id"] == "r1"
    assert by_id["n1"]["matched_review_id"] is None
    assert by_id["n2"]["attendees"] == ["정원석"]
    assert rev["matched_outing_id"] == "n2"


def test_match_orphan_review_far_date():
    notices = [_notice("n1", "2026-01-01", "인물")]
    rev = _review("r1", ["정원석"], datetime(2026, 6, 7), cat="인물", rod="2026-06-06")
    match_outings_with_reviews(notices + [rev])
    assert rev["matched_outing_id"] is None
    assert notices[0]["actually_held"] is False


def test_match_title_beats_category():
    """제목이 같은 짝을 놔두고 카테고리가 같은 남의 출사로 가면 안 된다.

    실제로 그랬다 — 카테고리 보너스가 100이라 나머지를 전부 눌렀다.
    """
    notices = [
        _notice("mine", "2026-03-04", "인물", "랩좀데니", "<03.04> (수)[인물]일본 가정집 코타츠"),
        _notice("other", "2026-03-08", "인물&풍경", "SUN", "3월 8일(일)[인풍] 광양매화마을"),
    ]
    rev = _review("r1", ["권두흥"], datetime(2026, 3, 4), cat="인물&풍경",
                  author="랩좀데니", rod="2026-03-04",
                  title="03.04(수) [인풍] 일본 가정집 코타츠 컨셉촬영")
    match_outings_with_reviews(notices + [rev])
    assert rev["matched_outing_id"] == "mine"


def test_match_does_not_swap_same_day_pair():
    """같은 사람이 같은 날짜로 올린 후기 둘이 서로 상대의 공지를 가져가지 않는다.

    후기 제목의 날짜가 하나 틀렸을 때(08.09 출사를 08.06으로 적음) 예전에는
    두 후기가 맞바뀌어 붙었다. 제목을 보고, 전체에서 확실한 짝부터 확정한다.
    """
    notices = [
        _notice("n_백범", "2025-08-06", "인물&풍경", "엄태진", "08.06(수) [인풍] 백범광장공원 성곽출사"),
        _notice("n_노들", "2025-08-09", "인물&풍경", "엄태진", "08.09(토)[인풍] 비오는 노들섬"),
    ]
    revs = [
        _review("r_노들", ["엄태진"], datetime(2025, 8, 10), cat="인물&풍경",
                author="엄태진", rod="2025-08-06", title="08.06(수)[인풍] 비오는노들섬 후기"),
        _review("r_백범", ["엄태진"], datetime(2025, 8, 10), cat="인물&풍경",
                author="엄태진", rod="2025-08-06", title="08.06(수)[인풍] 백범광장출사 후기"),
    ]
    match_outings_with_reviews(notices + revs)
    assert revs[0]["matched_outing_id"] == "n_노들"
    assert revs[1]["matched_outing_id"] == "n_백범"


def test_match_skips_canceled_notice():
    """펑 난 출사에는 후기가 없다. 붙으면 안 간 출사에 참석자가 얹힌다."""
    notices = [_notice("n1", "2026-02-21", "풍경", "정원석",
                       "(펑) 02.21(토) [풍경] 울산 통도사 매화출사", canceled=True)]
    rev = _review("r1", ["정원석"], datetime(2026, 2, 22), cat="풍경",
                  rod="2026-02-21", title="02.21 통도사 매화출사 후기")
    match_outings_with_reviews(notices + [rev])
    assert rev["matched_outing_id"] is None
    assert notices[0]["actually_held"] is False
    assert notices[0]["attendees"] == []


def test_match_needs_evidence_when_only_posted_date():
    """출사일을 못 읽었으면 작성일만으로는 안 붙인다 — 제목이나 작성자가 걸려야."""
    notices = [_notice("n1", "2026-04-11", "인물&풍경", "천성경", "04.11(토) [인풍] 올공 밤벚꽃")]
    rev = _review("r1", [], datetime(2026, 3, 31), cat="인물&풍경",
                  author="엄태진", title="[인풍] 반차 쓰고 벚꽃 출사")
    match_outings_with_reviews(notices + [rev])
    assert rev["matched_outing_id"] is None


def test_match_posted_date_ok_when_author_matches():
    notices = [_notice("n1", "2026-04-11", "인물&풍경", "엄태진", "04.11(토) [인풍] 올공 밤벚꽃")]
    rev = _review("r1", ["엄태진"], datetime(2026, 4, 12), cat="인물&풍경",
                  author="엄태진", title="[인풍] 올공 밤벚꽃 다녀왔습니다")
    match_outings_with_reviews(notices + [rev])
    assert rev["matched_outing_id"] == "n1"


def test_match_ignores_meetup_system_post():
    """소모임 정모 게시글은 후기가 아니다 — 공지를 선점하면 진짜 후기가 못 붙는다."""
    notices = [_notice("n1", "2026-04-02", "인물&풍경", "얀", "04.02 (목) [인풍] 벚꽃 나들이")]
    meetup = _review("m1", [], datetime(2026, 4, 1), cat="인물&풍경", author="얀",
                     title="[인풍] 벚꽃 나들이",
                     body="📌 정모 정보\n📅 4월 2일(목)\n📍 현충원\n💰 1/n")
    real = _review("r1", ["이하얀"], datetime(2026, 4, 3), cat="인물&풍경", author="얀",
                   rod="2026-04-02", title="04.02 (목) [인풍] 벚꽃 나들이 후기")
    match_outings_with_reviews(notices + [meetup, real])
    assert real["matched_outing_id"] == "n1"
    assert meetup["matched_outing_id"] is None
    assert meetup["is_meetup_post"] is True
    assert notices[0]["attendees"] == ["이하얀"]


def test_title_affinity_ignores_spacing_and_boilerplate():
    from core.collector import title_affinity
    # 붙여 쓰든 띄어 쓰든 같은 곳을 가리키는 것이 보여야 한다
    assert title_affinity("08.06(수)[인풍] 비오는노들섬 후기",
                          "08.09(토)[인풍] 비오는 노들섬") > 0.8
    # 딱지와 흔한 말("[인풍] … 출사 후기")만 겹치는 것은 닮은 게 아니다
    assert title_affinity("08.06(수)[인풍] 백범광장출사 후기",
                          "08.09(토)[인풍] 비오는 노들섬") == 0.0


# ── build_member_candidates (6차 마스터 editor 사전 채움) ────────
def test_candidates_body_frequency_to_real():
    body = "철수 영희 민수"
    posts = [{"cat": "E", "title": "", "body": body, "author": "닉a"} for _ in range(3)]
    rows = build_member_candidates(posts, [], min_freq=3)
    by_real = {r["실명"]: r for r in rows if r["실명"]}
    assert {"철수", "영희", "민수"} <= set(by_real)
    assert all(by_real[n]["포함"] for n in ("철수", "영희", "민수"))


def test_candidates_author_and_uploader_become_nick():
    posts = [{"cat": "A", "title": "", "body": "", "author": "닉a"}]
    photos = [{"author": "닉b"}, {"author": "닉a"}]
    rows = build_member_candidates(posts, photos, min_freq=99)
    by_nick = {r["닉네임"]: r for r in rows if r["닉네임"] and not r["실명"]}
    assert "닉a" in by_nick and "닉b" in by_nick


def test_candidates_blacklist_unchecked_but_present():
    # "후기"는 블랙리스트지만 작성자/업로더로 들어오면 행은 만들고 포함=False
    posts = [{"cat": "A", "title": "", "body": "", "author": "후기"}]
    rows = build_member_candidates(posts, [], min_freq=99)
    by_token = {(r["실명"] or r["닉네임"]): r for r in rows}
    assert "후기" in by_token
    assert by_token["후기"]["포함"] is False


def test_candidates_merge_same_token_across_sources():
    # 같은 토큰이 본문빈도(=실명)와 게시글 작성자(=닉네임) 둘 다 → 한 행에 둘 다 채움
    body = "정원석 정원석 정원석"
    posts = [{"cat": "E", "title": "", "body": body, "author": "정원석"}]
    rows = build_member_candidates(posts, [], min_freq=3)
    merged = [r for r in rows if r["실명"] == "정원석" and r["닉네임"] == "정원석"]
    assert len(merged) == 1


# ── 엑셀 번들 라운드트립 (no network) ────────────────────────────
def test_excel_multi_year_matrix_columns_grow():
    """다년 범위면 월 매트릭스 시트의 컬럼이 기간 길이만큼 늘어난다.

    예전에는 A~O 15열 고정이라 24개월을 표현할 수 없었다.
    """
    from io import BytesIO
    from openpyxl import load_workbook
    from core.excel_builder import build_excel

    blob = build_excel([], [], 202501, 202612)   # 24개월
    wb = load_workbook(BytesIO(blob))
    for sheet in ("🎨 월별 테마 매트릭스", "📅 월별 참석 매트릭스"):
        header = [c.value for c in wb[sheet][4]]
        assert len(header) == 2 + 24 + 1, (sheet, len(header))
        assert header[2] == "2025-01" and header[-2] == "2026-12"
        assert header[-1] in ("합계", "합계(장)")


def test_excel_single_month_matrix_does_not_break():
    """축이 1칸이어도 시트 생성이 깨지지 않아야 한다."""
    from io import BytesIO
    from openpyxl import load_workbook
    from core.excel_builder import build_excel

    blob = build_excel([], [], 202605, 202605)
    wb = load_workbook(BytesIO(blob))
    header = [c.value for c in wb["📅 월별 참석 매트릭스"][4]]
    assert len(header) == 2 + 1 + 1
    assert header[2] == "5월"


# ── v2: extract_raw_names / resolve_names / annotate_attendees ──
from core.collector import (
    LEFT_MEMBER, NOT_A_NAME,
    extract_raw_names, resolve_names, annotate_attendees, collect_all_unresolved,
)


def test_extract_raw_names_basic():
    raw = extract_raw_names("정원석 이하얀 김민수 후기", "후기")
    assert raw == ["정원석", "이하얀", "김민수"]  # "후기"는 블랙리스트 + 제목 strip


def test_extract_raw_names_english():
    raw = extract_raw_names("Daniel SUN 엄태진 함께", "")
    assert "Daniel" in raw and "SUN" in raw and "엄태진" in raw


def test_extract_raw_names_blacklist_english():
    raw = extract_raw_names("the and 엄태진 with", "")
    # "the","and","with"는 영문 블랙리스트로 제외
    assert raw == ["엄태진"]


def test_resolve_master_match():
    confirmed, unresolved = resolve_names(["엄태진", "승구"], {"승구", "엄태진"}, {})
    assert confirmed == ["엄태진", "승구"]
    assert unresolved == []


def test_resolve_nickname_mapping():
    confirmed, unresolved = resolve_names(["음승구"], {"승구"}, {"음승구": "승구"})
    assert confirmed == ["승구"]
    assert unresolved == []


def test_resolve_left_and_noise():
    confirmed, unresolved = resolve_names(
        ["엄태진", "민민기", "습니다"], {"엄태진"},
        {"민민기": LEFT_MEMBER, "습니다": NOT_A_NAME},
    )
    assert confirmed == ["엄태진"]
    assert unresolved == []


def test_resolve_unresolved():
    confirmed, unresolved = resolve_names(["엄태진", "처음본이름"], {"엄태진"}, {})
    assert confirmed == ["엄태진"]
    assert unresolved == ["처음본이름"]


def test_annotate_attendees_and_unresolved_counter():
    posts = [{
        "cat": "E", "title": "후기", "body": "정원석 이하얀 음승구 처음본이름",
        "author": "닉", "posted_at": datetime(2026, 6, 7),
    }]
    annotate_attendees(posts, {"정원석", "이하얀", "승구"}, {"음승구": "승구"})
    # 정규식이 한글 2~4자 → '처음본이름'(5자)은 토큰화 시 '처음본이'로 잘림
    assert posts[0]["attendees"] == ["정원석", "이하얀", "승구"]
    assert "처음본이" in posts[0]["unresolved_names"]
    cnt = collect_all_unresolved(posts)
    assert cnt["처음본이"] >= 1


# ── 9차: 가입인사 자동 매핑 + 동명이인 ─────────────────────────
from core.collector import (
    parse_join_name_aliases, find_duplicate_member_names,
)


def test_parse_join_name_aliases_basic():
    posts = [{
        "id": "j1", "author": "원석사진", "title": "가입인사",
        "body": "안녕하세요\n이름 : 정원석\n잘부탁드립니다",
        "posted_at": datetime(2025, 1, 1, 12, 0, 0),
    }]
    assert parse_join_name_aliases(posts) == {"정원석": "원석사진"}


def test_parse_join_name_aliases_self_name_skipped():
    # author == 본문 실명 → 매핑 의미 없음 → 제외
    posts = [{
        "id": "j1", "author": "정원석", "title": "",
        "body": "이름 : 정원석",
        "posted_at": datetime(2025, 1, 1),
    }]
    assert parse_join_name_aliases(posts) == {}


def test_parse_join_name_aliases_multiple_patterns():
    posts = [
        {"id": "1", "author": "닉a", "body": "성함 - 김민수", "posted_at": datetime(2025, 1, 1)},
        {"id": "2", "author": "닉b", "body": "본명 ：이하얀", "posted_at": datetime(2025, 1, 2)},
        {"id": "3", "author": "닉c", "body": "이름:박철수\n잘 부탁", "posted_at": datetime(2025, 1, 3)},
    ]
    out = parse_join_name_aliases(posts)
    assert out == {"김민수": "닉a", "이하얀": "닉b", "박철수": "닉c"}


def test_parse_join_name_aliases_skips_when_glued_to_hangul():
    # 정밀도 우선: '입니다'/'예요' 등 접미사가 공백 없이 붙으면 매칭 안 함 → 사용자가 보정
    posts = [{"id": "1", "author": "닉c", "body": "이름:박철수입니다",
              "posted_at": datetime(2025, 1, 1)}]
    assert parse_join_name_aliases(posts) == {}


def test_parse_join_name_aliases_latest_wins():
    # 같은 실명이 둘 다 잡히면 더 최근 글의 author로 덮어씀
    posts = [
        {"id": "1", "author": "옛닉", "body": "이름 : 정원석", "posted_at": datetime(2024, 1, 1)},
        {"id": "2", "author": "새닉", "body": "이름 : 정원석", "posted_at": datetime(2025, 6, 1)},
    ]
    assert parse_join_name_aliases(posts)["정원석"] == "새닉"


def test_parse_join_name_aliases_active_only():
    # active_mns로 필터 — 활성 멤버에 없는 닉네임의 글은 매핑 제외
    posts = [
        {"id": "1", "author": "탈퇴닉", "body": "이름 : 김민수", "posted_at": datetime(2024, 1, 1)},
        {"id": "2", "author": "활성닉", "body": "이름 : 이하얀", "posted_at": datetime(2025, 1, 1)},
    ]
    out = parse_join_name_aliases(posts, active_mns={"활성닉"})
    assert out == {"이하얀": "활성닉"}


def test_find_duplicate_member_names():
    members = [
        {"mid": "m1", "mn": "정원석"},
        {"mid": "m2", "mn": "정원석"},  # dup
        {"mid": "m3", "mn": "이하얀"},
        {"mid": "m4", "mn": "김민수"},
        {"mid": "m5", "mn": "김민수"},  # dup
        {"mid": "m6", "mn": ""},          # 빈 닉은 무시
    ]
    assert find_duplicate_member_names(members) == {"정원석", "김민수"}


def test_excel_listing_sheets_keep_year_in_month_column():
    """📝 후기글·📷 사진 시트의 '월' 열이 다년에서 연도를 잃지 않아야 한다.

    `.month`만 쓰면 2025-09와 2026-03이 9/3으로 남아 그 열로 정렬·필터할 때
    서로 다른 해가 섞인다. 앱의 같은 컬럼(reviews_table)과도 표기를 맞춘다.
    """
    from io import BytesIO
    from openpyxl import load_workbook
    from core.excel_builder import build_excel

    def review(pid, posted):
        return {"id": pid, "author": "닉", "wid": "w1", "title": f"{pid} 후기",
                "body": "", "outing_date": None, "posted_at": posted,
                "cat": "E", "cat_label": "후기", "category": "풍경",
                "is_outing": False, "is_canceled": False,
                "likes": 0, "comments": 0, "images": 0,
                "needs_review": False, "review_reason": ""}

    def photo(pid, posted):
        return {"id": pid, "author": "닉", "wid": "w1", "posted_at": posted,
                "likes": 0, "comments": 0, "has_comment": False,
                "url_large": "x", "url_medium": "y", "url_small": "z", "url_thumb": "n"}

    posts = [review("r1", datetime(2025, 9, 6, 9, 0)),
             review("r2", datetime(2026, 3, 8, 9, 0))]
    photos = [photo("p1", datetime(2025, 9, 7, 9, 0)),
              photo("p2", datetime(2026, 3, 9, 9, 0))]

    wb = load_workbook(BytesIO(build_excel(posts, photos, 202509, 202603)))
    for sheet in ("📝 후기글", "📷 사진"):
        ws = wb[sheet]
        header = [c.value for c in ws[2]]
        col = header.index("월") + 1
        months = {ws.cell(r, col).value for r in range(3, ws.max_row + 1)}
        assert months == {"2025-09", "2026-03"}, (sheet, months)
