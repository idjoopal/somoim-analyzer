"""
다감노📸 엑셀 빌더

수집된 게시글·사진 데이터를 받아 다중 시트 엑셀(bytes)을 생성.

외부 의존: openpyxl

주요 함수:
- build_excel(posts, photos, year, month=None) -> bytes
- save_excel(posts, photos, year, month=None, path=...) -> str
"""

from __future__ import annotations

from io import BytesIO
from datetime import datetime, date
from collections import defaultdict
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

from .collector import GROUP_NAME, OUTING_CATS, NON_OUTING_CATS, find_duplicate_member_names  # noqa: F401


# ═══════════════════════════════════════════════════════════════
# 엑셀 = 결과 + 원본 (다음 세션 재사용)
# ═══════════════════════════════════════════════════════════════
#
# 11개 가시 시트 뒤에 4개 숨김 시트를 둬서, 이 엑셀 하나만으로
# (인사이트 다시 보기) + (원본 데이터 재분석)이 가능하게 한다.
# 숨김 시트는 Excel에서 시트 탭 우클릭 → 숨기기 취소로 노출 가능.

EXCEL_SCHEMA_VERSION = 1

BASE_POST_KEYS: list[str] = [
    "id", "author", "wid", "title", "body", "outing_date", "posted_at",
    "cat", "cat_label", "category", "is_outing", "is_canceled",
    "likes", "comments", "images", "needs_review", "review_reason",
]
BASE_PHOTO_KEYS: list[str] = [
    "id", "author", "wid", "posted_at", "likes", "comments", "has_comment",
    "url_large", "url_medium", "url_small", "url_thumb",
]
BASE_MEMBER_KEYS: list[str] = [
    "mid", "mn", "is_admin", "joined_at", "last_visit", "os", "push",
]
_BOOL_POST_KEYS = {"is_outing", "is_canceled", "needs_review"}
_BOOL_PHOTO_KEYS = {"has_comment"}
_BOOL_MEMBER_KEYS = {"is_admin", "push"}
_DT_MEMBER_KEYS = {"joined_at", "last_visit"}
_CELL_MAX_LEN = 32000  # 엑셀 셀당 32,767자 한도 안전 여유


# ═══════════════════════════════════════════════════════════════
# 스타일
# ═══════════════════════════════════════════════════════════════

C = dict(
    HDR_DARK="1F3864", HDR_MID="2E75B6", HDR_LIGHT="BDD7EE",
    ACCENT_RED="C00000", ACCENT_GRN="375623",
    ACCENT_YLW="FFC000", ACCENT_PRP="7030A0",
    OUTING="E2EFDA", CANCEL="FCE4D6", REVIEW="EBF3FB",
    PHOTO="FFF2CC", THEME="F2DCDB",
    WHITE="FFFFFF", GRAY_LIGHT="F5F5F5",
)


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_): return PatternFill("solid", fgColor=hex_)
def _center():   return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left():     return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _hdr_font(size=11, white=True):
    return Font(name="Arial", bold=True, size=size,
                color=C["WHITE"] if white else C["HDR_DARK"])

def _body_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _style_header_row(ws, row, c0, c1, bg=None):
    bg = bg or C["HDR_DARK"]
    for c in range(c0, c1 + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _hdr_font()
        cell.fill = _fill(bg)
        cell.alignment = _center()
        cell.border = _thin_border()

def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _title_band(ws, text, cols, row=1, height=36, bg=None, size=14):
    bg = bg or C["HDR_DARK"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = _hdr_font(size, True)
    cell.fill = _fill(bg)
    cell.alignment = _center()
    ws.row_dimensions[row].height = height


# ═══════════════════════════════════════════════════════════════
# 메인 빌더
# ═══════════════════════════════════════════════════════════════

def build_excel(
    posts: list[dict],
    photos: list[dict],
    year: int,
    month: Optional[int] = None,
    master_records: Optional[list[dict]] = None,
    members: Optional[list[dict]] = None,
    banned: Optional[set[str]] = None,
    resolution: Optional[dict[str, str]] = None,
    join_aliases: Optional[dict[str, str]] = None,
) -> bytes:
    """
    수집 데이터로부터 엑셀 파일(bytes) 생성.

    가시 시트 11개 + 숨김 시트 3~4개(원본 + 마스터 임베드 — 재업로드 시 API 호출 없이 재분석).

    시트 구성:
    1. 📊 대시보드
    2. 👤 게시글 통계
    3. 📌 출사 공지
    4. 📝 후기글
    5. 📷 사진
    6. 🎨 월별 테마 매트릭스
    7. 👤 사진 통계
    8. 💡 인사이트
    9. 🎯 출사별 참석자       — annotate_review_attendees + match_outings_with_reviews 거친 경우
    10. 👥 멤버별 참석
    11. 📅 월별 참석 매트릭스
    (숨김) _메타, _원본_게시글, _원본_사진, _멤버마스터(master_records 있을 때만)
    """
    period_label = f"{year}년" + (f" {month}월" if month else "")

    posts_A = [p for p in posts if p["cat"] == "A"]
    posts_E = [p for p in posts if p["cat"] == "E"]
    posts_J = [p for p in posts if p["cat"] == "J"]
    posts_canceled = [p for p in posts_A if p["is_canceled"]]
    posts_active   = [p for p in posts_A if not p["is_canceled"]]
    photos_with_cmt = [p for p in photos if p["has_comment"]]

    wb = Workbook()
    wb.remove(wb.active)

    _build_sheet_dashboard(wb, posts, posts_A, posts_E, posts_active,
                           posts_canceled, photos, photos_with_cmt,
                           year, month, period_label)
    user_stats, sorted_users = _build_sheet_post_stats(wb, posts)
    _build_sheet_outings(wb, posts_A)
    _build_sheet_reviews(wb, posts_E)
    _build_sheet_photos(wb, photos)
    user_month, mon_user_count, sorted_authors = _build_sheet_theme_matrix(wb, photos, photos_with_cmt)
    photo_stats, sorted_photo_users = _build_sheet_photo_stats(wb, photos)
    _build_sheet_insights(wb, posts, posts_A, posts_E, posts_active,
                          posts_canceled, photos, photos_with_cmt,
                          user_stats, sorted_users,
                          user_month, mon_user_count, sorted_authors,
                          photo_stats, sorted_photo_users,
                          period_label)
    _build_sheet_outing_attendees(wb, posts_A)
    _build_sheet_member_attendance(wb, posts_A)
    _build_sheet_monthly_matrix(wb, posts_A)
    if members:
        _build_sheet_member_overview(wb, members, posts, photos, posts_A)

    _build_sheet_meta(wb, year, month)
    _build_sheet_raw_posts(wb, posts)
    _build_sheet_raw_photos(wb, photos)
    if master_records is not None:
        _build_sheet_master(wb, master_records)
    if members is not None:
        _build_sheet_members(wb, members)
    if banned:
        _build_sheet_banned(wb, banned)
    if resolution is not None:
        _build_sheet_resolution(wb, resolution)
    if join_aliases is not None:
        _build_sheet_join_aliases(wb, join_aliases)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_excel(
    posts: list[dict],
    photos: list[dict],
    year: int,
    month: Optional[int] = None,
    path: Optional[str] = None,
    master_records: Optional[list[dict]] = None,
    members: Optional[list[dict]] = None,
    banned: Optional[set[str]] = None,
    resolution: Optional[dict[str, str]] = None,
    join_aliases: Optional[dict[str, str]] = None,
) -> str:
    """엑셀을 파일로 저장. 경로 미지정시 기본 이름 사용."""
    if path is None:
        period = f"{year}" + (f"_{month:02d}" if month else "")
        path = f"다감노_{period}_분석.xlsx"
    data = build_excel(posts, photos, year, month,
                       master_records=master_records,
                       members=members, banned=banned, resolution=resolution,
                       join_aliases=join_aliases)
    with open(path, "wb") as f:
        f.write(data)
    return path


# ═══════════════════════════════════════════════════════════════
# 숨김 시트 — 원본 데이터 + 마스터 임베드
# ═══════════════════════════════════════════════════════════════

def _truncate(s: str, limit: int = _CELL_MAX_LEN) -> str:
    return s if len(s) <= limit else s[:limit]


def _to_cell(value, key: str = "") -> object:
    """원본 dict 값을 엑셀 셀에 쓰기 좋은 형태로 정규화."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, datetime, date)):
        return value
    if isinstance(value, str):
        return _truncate(value) if key == "body" else value
    return str(value)


def _build_sheet_meta(wb: Workbook, year: int, month: Optional[int]) -> None:
    ws = wb.create_sheet("_메타")
    ws.sheet_state = "hidden"
    ws.append(["key", "value"])
    rows = [
        ("schema_version", EXCEL_SCHEMA_VERSION),
        ("year", int(year)),
        ("month", "" if month is None else int(month)),
        ("saved_at", datetime.now().isoformat(timespec="seconds")),
    ]
    for k, v in rows:
        ws.append([k, v])


def _build_sheet_raw_posts(wb: Workbook, posts: list[dict]) -> None:
    ws = wb.create_sheet("_원본_게시글")
    ws.sheet_state = "hidden"
    ws.append(BASE_POST_KEYS)
    for p in posts:
        ws.append([_to_cell(p.get(k), k) for k in BASE_POST_KEYS])


def _build_sheet_raw_photos(wb: Workbook, photos: list[dict]) -> None:
    ws = wb.create_sheet("_원본_사진")
    ws.sheet_state = "hidden"
    ws.append(BASE_PHOTO_KEYS)
    for ph in photos:
        ws.append([_to_cell(ph.get(k), k) for k in BASE_PHOTO_KEYS])


def _build_sheet_master(wb: Workbook, master_records: list[dict]) -> None:
    ws = wb.create_sheet("_멤버마스터")
    ws.sheet_state = "hidden"
    ws.append(["실명", "닉네임", "별칭"])
    for r in master_records or []:
        aliases = r.get("별칭") or []
        if isinstance(aliases, str):
            alias_str = aliases
        else:
            alias_str = ";".join(str(a) for a in aliases if a)
        ws.append([
            str(r.get("실명") or ""),
            str(r.get("닉네임") or ""),
            alias_str,
        ])


def _build_sheet_members(wb: Workbook, members: list[dict]) -> None:
    """활성 멤버 API 결과를 숨김 시트로 저장(엑셀 재업로드 시 복원용)."""
    ws = wb.create_sheet("_멤버")
    ws.sheet_state = "hidden"
    ws.append(BASE_MEMBER_KEYS)
    for m in members or []:
        ws.append([_to_cell(m.get(k), k) for k in BASE_MEMBER_KEYS])


def _build_sheet_banned(wb: Workbook, banned: set[str]) -> None:
    """탈퇴 멤버 닉네임 1컬럼."""
    ws = wb.create_sheet("_탈퇴멤버")
    ws.sheet_state = "hidden"
    ws.append(["닉네임"])
    for nick in sorted(banned or set()):
        ws.append([str(nick)])


def _build_sheet_resolution(wb: Workbook, resolution: dict[str, str]) -> None:
    """이름 해소 매핑(이름→처리) 저장."""
    ws = wb.create_sheet("_이름매핑")
    ws.sheet_state = "hidden"
    ws.append(["이름", "처리"])
    for name, target in (resolution or {}).items():
        ws.append([str(name), str(target)])


def _build_sheet_join_aliases(wb: Workbook, join_aliases: dict[str, str]) -> None:
    """가입인사 자동 추출 매핑(실명→닉네임) 저장."""
    ws = wb.create_sheet("_가입인사매핑")
    ws.sheet_state = "hidden"
    ws.append(["실명", "닉네임"])
    for real, nick in (join_aliases or {}).items():
        ws.append([str(real), str(nick)])


def _build_sheet_member_overview(
    wb: Workbook, members: list[dict],
    posts: list[dict], photos: list[dict], posts_A: list[dict],
) -> None:
    """🧑‍🤝‍🧑 멤버 현황 (가시 시트). 활동 집계 + 활동 상태 분류."""
    from datetime import datetime as _dt
    ws = wb.create_sheet("🧑‍🤝‍🧑 멤버 현황")
    cols = 9
    _title_band(ws, "🧑‍🤝‍🧑 멤버 현황 — 활성 멤버 활동 분포", cols)
    headers = ["닉네임", "운영진", "가입일", "마지막 방문", "OS",
               "게시글", "사진", "참석", "활동상태"]
    ws.append(headers)
    _style_header_row(ws, 2, 1, cols)

    post_count: dict[str, int] = defaultdict(int)
    for p in posts:
        if p.get("author"):
            post_count[p["author"]] += 1
    photo_count: dict[str, int] = defaultdict(int)
    for ph in photos:
        if ph.get("author"):
            photo_count[ph["author"]] += 1
    attendance_count: dict[str, int] = defaultdict(int)
    for a in posts_A:
        for n in a.get("attendees", []) or []:
            attendance_count[n] += 1

    now = _dt.now()
    rows: list[tuple] = []
    for m in members:
        mn = m.get("mn", "")
        if not mn:
            continue
        last = m.get("last_visit")
        days_idle = (now - last).days if last else 9999
        pcnt = post_count.get(mn, 0)
        phcnt = photo_count.get(mn, 0)
        acnt = attendance_count.get(mn, 0)
        if pcnt == 0 and phcnt == 0 and acnt == 0:
            status = "유령"
        elif days_idle >= 90:
            status = "휴면"
        else:
            status = "활성"
        rows.append((mn,
                     "Y" if m.get("is_admin") else "",
                     last.date() if m.get("joined_at") else None,
                     last.date() if last else None,
                     m.get("os") or "",
                     pcnt, phcnt, acnt, status))
        # NOTE: 가입일은 별도로 다시 계산해야 함 — 위 last를 잘못 썼음. 아래에서 정정.

    # 가입일을 정확히 다시 계산 (위 임시 코드 정정용)
    rows = []
    for m in members:
        mn = m.get("mn", "")
        if not mn:
            continue
        joined = m.get("joined_at")
        last = m.get("last_visit")
        days_idle = (now - last).days if last else 9999
        pcnt = post_count.get(mn, 0)
        phcnt = photo_count.get(mn, 0)
        acnt = attendance_count.get(mn, 0)
        if pcnt == 0 and phcnt == 0 and acnt == 0:
            status = "유령"
        elif days_idle >= 90:
            status = "휴면"
        else:
            status = "활성"
        rows.append((mn,
                     "Y" if m.get("is_admin") else "",
                     joined.date() if joined else None,
                     last.date() if last else None,
                     m.get("os") or "",
                     pcnt, phcnt, acnt, status))

    # 활동상태(활성→휴면→유령) · 활동량 desc
    status_order = {"활성": 0, "휴면": 1, "유령": 2}
    rows.sort(key=lambda x: (status_order.get(x[8], 9),
                              -(x[5] + x[6] + x[7])))

    duplicates = find_duplicate_member_names(members)
    for r in rows:
        ws.append(list(r))
        last_row = ws.max_row
        status = r[8]
        bg = C["OUTING"] if status == "활성" else (C["GRAY_LIGHT"] if status == "휴면" else C["CANCEL"])
        for c in range(1, cols + 1):
            ws.cell(row=last_row, column=c).border = _thin_border()
        ws.cell(row=last_row, column=cols).fill = _fill(bg)
        for c in (3, 4):
            cell = ws.cell(row=last_row, column=c)
            if cell.value:
                cell.number_format = "yyyy-mm-dd"
        if r[0] in duplicates:
            nick_cell = ws.cell(row=last_row, column=1)
            nick_cell.value = f"⚠️ {r[0]}"
            nick_cell.fill = _fill(C["ACCENT_YLW"])
            nick_cell.font = Font(name="Arial", size=10, bold=True, color=C["ACCENT_RED"])

    _set_col_widths(ws, {"A": 16, "B": 8, "C": 12, "D": 12, "E": 10,
                          "F": 8, "G": 8, "H": 8, "I": 10})


# ═══════════════════════════════════════════════════════════════
# 엑셀 → 원본 데이터 복원 (재업로드)
# ═══════════════════════════════════════════════════════════════

def load_excel_bundle(data: bytes) -> dict:
    """엑셀 bytes에서 숨김 시트의 원본 데이터를 복원.

    Returns:
        {
          "year": int, "month": int|None,
          "posts":  list[dict],
          "photos": list[dict],
          "master": list[dict],     # legacy(_멤버마스터) — 비어있을 수 있음
          "members": list[dict],    # _멤버 (활성 멤버 API 결과)
          "banned": set[str],       # _탈퇴멤버
          "resolution": dict[str,str],  # _이름매핑 (이름→처리)
        }
    Raises:
        ValueError — 메타 시트 없음/버전 불일치/필수 시트 누락.
    """
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(data), data_only=True)

    for name in ("_메타", "_원본_게시글", "_원본_사진"):
        if name not in wb.sheetnames:
            raise ValueError(f"이 엑셀에는 원본 데이터({name})가 포함돼 있지 않습니다.")

    meta = _read_meta(wb["_메타"])
    if meta.get("schema_version") != EXCEL_SCHEMA_VERSION:
        raise ValueError(
            f"지원하지 않는 엑셀 버전: {meta.get('schema_version')} "
            f"(지원={EXCEL_SCHEMA_VERSION})"
        )
    year = int(meta["year"])
    month_val = meta.get("month")
    month: Optional[int] = int(month_val) if month_val not in (None, "", 0) else None

    posts = _read_raw(wb["_원본_게시글"], BASE_POST_KEYS, _BOOL_POST_KEYS)
    photos = _read_raw(wb["_원본_사진"], BASE_PHOTO_KEYS, _BOOL_PHOTO_KEYS)
    master = _read_master(wb["_멤버마스터"]) if "_멤버마스터" in wb.sheetnames else []
    members = _read_members(wb["_멤버"]) if "_멤버" in wb.sheetnames else []
    banned = _read_banned(wb["_탈퇴멤버"]) if "_탈퇴멤버" in wb.sheetnames else set()
    resolution = _read_resolution(wb["_이름매핑"]) if "_이름매핑" in wb.sheetnames else {}
    join_aliases = (_read_join_aliases(wb["_가입인사매핑"])
                    if "_가입인사매핑" in wb.sheetnames else {})

    return {"year": year, "month": month, "posts": posts, "photos": photos,
            "master": master, "members": members, "banned": banned,
            "resolution": resolution, "join_aliases": join_aliases}


def _read_meta(ws) -> dict:
    out: dict = {}
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return out
    for row in rows[1:]:  # 헤더 스킵
        if not row or row[0] in (None, ""):
            continue
        out[str(row[0])] = row[1] if len(row) > 1 else None
    return out


def _read_raw(ws, keys: list[str], bool_keys: set[str]) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c) if c is not None else "" for c in rows[0]]
    out: list[dict] = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        rec: dict = {}
        for i, k in enumerate(header):
            v = row[i] if i < len(row) else None
            if k in bool_keys:
                v = bool(v)
            elif k == "review_reason" and v is None:
                v = ""
            rec[k] = v
        for k in keys:
            rec.setdefault(k, "" if k == "review_reason" else None)
        out.append(rec)
    return out


def _read_master(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    out: list[dict] = []
    for row in rows[1:]:
        if not row or all(c in (None, "") for c in row):
            continue
        real = str(row[0]) if len(row) > 0 and row[0] not in (None, "") else ""
        nick = str(row[1]) if len(row) > 1 and row[1] not in (None, "") else ""
        alias_raw = str(row[2]) if len(row) > 2 and row[2] not in (None, "") else ""
        aliases = [a.strip() for a in alias_raw.split(";") if a.strip()]
        out.append({"실명": real, "닉네임": nick, "별칭": aliases})
    return out


def _read_members(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c) if c is not None else "" for c in rows[0]]
    out: list[dict] = []
    for row in rows[1:]:
        if not row or all(c in (None, "") for c in row):
            continue
        rec: dict = {}
        for i, k in enumerate(header):
            v = row[i] if i < len(row) else None
            if k in _BOOL_MEMBER_KEYS:
                v = bool(v)
            elif k in _DT_MEMBER_KEYS:
                if v in (None, ""):
                    v = None
            elif v is None:
                v = ""
            rec[k] = v
        for k in BASE_MEMBER_KEYS:
            rec.setdefault(k, None)
        out.append(rec)
    return out


def _read_banned(ws) -> set[str]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return set()
    out: set[str] = set()
    for row in rows[1:]:
        if row and row[0] not in (None, ""):
            out.add(str(row[0]))
    return out


def _read_resolution(ws) -> dict[str, str]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    out: dict[str, str] = {}
    for row in rows[1:]:
        if not row or row[0] in (None, ""):
            continue
        name = str(row[0])
        target = str(row[1]) if len(row) > 1 and row[1] not in (None, "") else ""
        if target:
            out[name] = target
    return out


def _read_join_aliases(ws) -> dict[str, str]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    out: dict[str, str] = {}
    for row in rows[1:]:
        if not row or row[0] in (None, ""):
            continue
        real = str(row[0])
        nick = str(row[1]) if len(row) > 1 and row[1] not in (None, "") else ""
        if nick:
            out[real] = nick
    return out


# ═══════════════════════════════════════════════════════════════
# 시트 빌더들
# ═══════════════════════════════════════════════════════════════

def _build_sheet_dashboard(wb, posts, posts_A, posts_E, posts_active,
                           posts_canceled, photos, photos_with_cmt,
                           year, month, period_label):
    ws = wb.create_sheet("📊 대시보드")
    ws.sheet_view.showGridLines = False

    _title_band(ws, f"📸 {GROUP_NAME} {period_label} 활동 대시보드", 12, height=40, size=16)
    ws.merge_cells("A2:L2")
    ws["A2"] = f"추출일: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  기간: {period_label}"
    ws["A2"].font = Font(name="Arial", size=10, color="888888")
    ws["A2"].alignment = _center()
    ws.row_dimensions[2].height = 18

    # KPI 카드
    kpis = [
        ("전체 게시글", len(posts),          C["HDR_DARK"],   "A"),
        ("진행 출사",  len(posts_active),    C["ACCENT_GRN"], "C"),
        ("취소 출사",  len(posts_canceled),  C["ACCENT_RED"], "E"),
        ("후기글",     len(posts_E),         "2E75B6",        "G"),
        ("사진 업로드", len(photos),          "ED7D31",        "I"),
        ("테마 예상",  len(photos_with_cmt), C["ACCENT_PRP"], "K"),
    ]
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 36
    ws.row_dimensions[6].height = 6
    for label, val, color, col in kpis:
        c_idx = ord(col) - ord("A") + 1
        ws.merge_cells(start_row=4, start_column=c_idx, end_row=4, end_column=c_idx+1)
        ws.merge_cells(start_row=5, start_column=c_idx, end_row=5, end_column=c_idx+1)
        for r in (4, 5):
            for cc in (c_idx, c_idx+1):
                ws.cell(r, cc).fill = _fill(color)
                ws.cell(r, cc).border = _thin_border()
        ws.cell(4, c_idx).value = label
        ws.cell(4, c_idx).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ws.cell(4, c_idx).alignment = _center()
        ws.cell(5, c_idx).value = val
        ws.cell(5, c_idx).font = Font(name="Arial", bold=True, size=22, color="FFFFFF")
        ws.cell(5, c_idx).alignment = _center()

    # 월별 추이
    ws.merge_cells("A8:L8")
    ws["A8"] = "📅 월별 활동 현황"
    ws["A8"].font = Font(name="Arial", bold=True, size=12)
    ws["A8"].fill = _fill(C["HDR_LIGHT"])
    ws["A8"].alignment = _left()
    ws.row_dimensions[8].height = 22
    ws.row_dimensions[9].height = 6

    row_h = 10
    ws.cell(row_h, 1).value = "구분"
    for i, m in enumerate(range(1, 13)):
        ws.cell(row_h, i+2).value = f"{m}월"
    _style_header_row(ws, row_h, 1, 13)
    ws.row_dimensions[row_h].height = 22

    def mon_outing(items, m):
        return sum(1 for x in items
                   if x["outing_date"] and date.fromisoformat(x["outing_date"]).month == m)

    row_data = [
        ("진행 출사", [mon_outing(posts_active, m)   for m in range(1, 13)], C["ACCENT_GRN"]),
        ("취소 출사", [mon_outing(posts_canceled, m) for m in range(1, 13)], C["ACCENT_RED"]),
        ("후기글",   [sum(1 for p in posts_E if p["posted_at"].month == m) for m in range(1, 13)], "2E75B6"),
        ("사진",     [sum(1 for p in photos  if p["posted_at"].month == m) for m in range(1, 13)], "ED7D31"),
        ("테마 예상", [sum(1 for p in photos_with_cmt if p["posted_at"].month == m) for m in range(1, 13)], C["ACCENT_PRP"]),
    ]
    for i, (label, vals, bg) in enumerate(row_data):
        r = row_h + 1 + i
        ws.row_dimensions[r].height = 20
        ws.cell(r, 1).value = label
        ws.cell(r, 1).fill = _fill(bg)
        ws.cell(r, 1).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ws.cell(r, 1).alignment = _center()
        ws.cell(r, 1).border = _thin_border()
        for j, v in enumerate(vals):
            cell = ws.cell(r, j+2, value=v)
            cell.alignment = _center()
            cell.border = _thin_border()
            if v > 0:
                cell.fill = _fill(bg)
                cell.font = Font(name="Arial", size=10, color="FFFFFF", bold=True)

    # 차트
    chart = BarChart()
    chart.type = "col"; chart.grouping = "clustered"
    chart.title = "월별 출사 공지 (진행/취소)"
    chart.y_axis.title = "건수"; chart.x_axis.title = "월"
    chart.style = 10; chart.width = 22; chart.height = 11
    cats_ref = Reference(ws, min_col=2, max_col=13, min_row=row_h, max_row=row_h)
    for i, label in enumerate(["진행 출사", "취소 출사"], 1):
        data = Reference(ws, min_col=2, max_col=13, min_row=row_h+i, max_row=row_h+i)
        chart.add_data(data)
        chart.series[-1].title = SeriesLabel(v=label)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "A17")

    _set_col_widths(ws, {"A": 14, **{get_column_letter(i): 6 for i in range(2, 14)}})


def _build_sheet_post_stats(wb, posts):
    ws = wb.create_sheet("👤 게시글 통계")
    ws.sheet_view.showGridLines = False
    _title_band(ws, "👤 사용자별 게시글 활동 통계", 14)

    user_stats: dict = defaultdict(lambda: {
        "A": 0, "A_취소": 0, "E": 0,
        "cats": defaultdict(int),
        "likes": 0, "comments": 0,
    })
    for p in posts:
        s = user_stats[p["author"]]
        if p["cat"] == "A":
            if p["is_canceled"]: s["A_취소"] += 1
            else:                s["A"]       += 1
        elif p["cat"] == "E":
            s["E"] += 1
        if p["category"]:
            s["cats"][p["category"]] += 1
        s["likes"]    += p["likes"]
        s["comments"] += p["comments"]

    def utotal(s): return s["A"] + s["A_취소"] + s["E"]
    sorted_users = sorted(user_stats.items(), key=lambda x: -utotal(x[1]))

    hdrs = ["순위","작성자","공지","취소","후기","합계","인물","인물&풍경","풍경",
            "보정","GN","문화","취소율","좋아요"]
    ws.append([])
    ws.append(hdrs)
    ws.row_dimensions[2].height = 6
    ws.row_dimensions[3].height = 24
    _style_header_row(ws, 3, 1, len(hdrs))
    ws.freeze_panes = "C4"

    for rank, (author, s) in enumerate(sorted_users, 1):
        notice = s["A"] + s["A_취소"]
        rate = f"{s['A_취소']/notice*100:.1f}%" if notice else "-"
        r = ws.max_row + 1
        ws.append([
            rank, author, s["A"], s["A_취소"], s["E"], utotal(s),
            s["cats"].get("인물", 0), s["cats"].get("인물&풍경", 0), s["cats"].get("풍경", 0),
            s["cats"].get("보정", 0), s["cats"].get("GN", 0), s["cats"].get("문화", 0),
            rate, s["likes"],
        ])
        bg = C["GRAY_LIGHT"] if rank % 2 == 0 else C["WHITE"]
        for c in range(1, len(hdrs)+1):
            cell = ws.cell(r, c)
            cell.font = _body_font(bold=(c == 2))
            cell.alignment = _left() if c == 2 else _center()
            cell.border = _thin_border()
            cell.fill = _fill(bg)
        if notice >= 3:
            pct = s["A_취소"] / notice * 100
            if pct >= 50:
                ws.cell(r, 13).font = Font(name="Arial", size=10, bold=True, color=C["ACCENT_RED"])
            elif pct >= 25:
                ws.cell(r, 13).font = Font(name="Arial", size=10, bold=True, color="E26B0A")
        ws.row_dimensions[r].height = 20

    n = len(sorted_users)
    if n > 0:
        ws.conditional_formatting.add(
            f"F4:F{3+n}",
            DataBarRule(start_type="min", end_type="max", color="2E75B6", showValue=True),
        )
    _set_col_widths(ws, {
        "A":5,"B":12,"C":6,"D":6,"E":6,"F":6,"G":6,"H":11,"I":6,
        "J":6,"K":6,"L":6,"M":8,"N":7,
    })

    return user_stats, sorted_users


def _build_sheet_outings(wb, posts_A):
    ws = wb.create_sheet("📌 출사 공지")
    ws.sheet_view.showGridLines = False
    _title_band(ws, "📌 출사 공지 전체 (cat=A) — 출사일 기준 정렬", 10, height=32, size=13)

    hdrs = ["출사날짜","공지날짜","D-day","작성자","카테고리","유형","상태","제목","좋아요","댓글"]
    ws.append(hdrs)
    _style_header_row(ws, 2, 1, len(hdrs), bg=C["HDR_MID"])
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    for p in sorted(posts_A, key=lambda x: x["outing_date"] or "0000", reverse=True):
        od  = p["outing_date"] or "-"
        pd_ = p["posted_at"].strftime("%Y-%m-%d")
        if p["outing_date"]:
            dday = (date.fromisoformat(p["outing_date"]) - p["posted_at"].date()).days
            dday_str = f"+{dday}일" if dday >= 0 else f"{dday}일"
        else:
            dday_str = "-"
        r = ws.max_row + 1
        ws.append([
            od, pd_, dday_str, p["author"],
            p["category"] or "-",
            "출사" if p["is_outing"] else "활동",
            "취소" if p["is_canceled"] else "진행",
            p["title"], p["likes"], p["comments"],
        ])
        bg = C["CANCEL"] if p["is_canceled"] else C["OUTING"]
        for c in range(1, len(hdrs)+1):
            cell = ws.cell(r, c)
            cell.font = _body_font()
            cell.fill = _fill(bg)
            cell.border = _thin_border()
            cell.alignment = _left() if c == 8 else _center()
        if p["is_canceled"]:
            ws.cell(r, 7).font = Font(name="Arial", size=10, bold=True, color=C["ACCENT_RED"])
        ws.row_dimensions[r].height = 18

    _set_col_widths(ws, {"A":12,"B":12,"C":9,"D":10,"E":9,"F":7,"G":7,"H":45,"I":7,"J":7})


def _build_sheet_reviews(wb, posts_E):
    ws = wb.create_sheet("📝 후기글")
    ws.sheet_view.showGridLines = False
    _title_band(ws, "📝 후기 게시글 (cat=E)", 8, height=32, size=13)

    hdrs = ["작성일","월","작성자","카테고리","유형","제목","좋아요","댓글"]
    ws.append(hdrs)
    _style_header_row(ws, 2, 1, len(hdrs), bg=C["HDR_MID"])
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    for p in sorted(posts_E, key=lambda x: x["posted_at"], reverse=True):
        r = ws.max_row + 1
        ws.append([
            p["posted_at"].strftime("%Y-%m-%d"), p["posted_at"].month,
            p["author"], p["category"] or "-",
            "출사" if p["is_outing"] else "활동",
            p["title"], p["likes"], p["comments"],
        ])
        bg = C["REVIEW"] if r % 2 == 0 else C["WHITE"]
        for c in range(1, len(hdrs)+1):
            cell = ws.cell(r, c)
            cell.font = _body_font()
            cell.fill = _fill(bg)
            cell.border = _thin_border()
            cell.alignment = _left() if c == 6 else _center()
        ws.row_dimensions[r].height = 18

    _set_col_widths(ws, {"A":12,"B":5,"C":10,"D":9,"E":7,"F":48,"G":7,"H":7})


def _build_sheet_photos(wb, photos):
    ws = wb.create_sheet("📷 사진")
    ws.sheet_view.showGridLines = False
    _title_band(ws, "📷 사진 업로드 (테마 참여 예상 = 댓글 받은 사진)", 9, height=32, size=13)

    hdrs = ["업로드일","월","작성자","좋아요","댓글","테마예상","사진ID","고화질 URL","썸네일 URL"]
    ws.append(hdrs)
    _style_header_row(ws, 2, 1, len(hdrs), bg=C["HDR_MID"])
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    for p in sorted(photos, key=lambda x: x["posted_at"], reverse=True):
        r = ws.max_row + 1
        ws.append([
            p["posted_at"].strftime("%Y-%m-%d %H:%M"),
            p["posted_at"].month,
            p["author"], p["likes"], p["comments"],
            "🎨 예상" if p["has_comment"] else "",
            p["id"], p["url_large"], p["url_thumb"],
        ])
        bg = C["THEME"] if p["has_comment"] else (C["GRAY_LIGHT"] if r % 2 == 0 else C["PHOTO"])
        for c in range(1, len(hdrs)+1):
            cell = ws.cell(r, c)
            cell.font = _body_font()
            cell.fill = _fill(bg)
            cell.border = _thin_border()
            cell.alignment = _left() if c in (7, 8, 9) else _center()
        if p["has_comment"]:
            ws.cell(r, 6).font = Font(name="Arial", size=10, bold=True, color=C["ACCENT_PRP"])
        ws.cell(r, 8).hyperlink = p["url_large"]
        ws.cell(r, 8).font = Font(name="Arial", size=9, color="0563C1", underline="single")
        ws.cell(r, 9).hyperlink = p["url_thumb"]
        ws.cell(r, 9).font = Font(name="Arial", size=9, color="0563C1", underline="single")
        ws.row_dimensions[r].height = 18

    _set_col_widths(ws, {"A":16,"B":5,"C":10,"D":7,"E":7,"F":9,"G":40,"H":62,"I":62})


def _build_sheet_theme_matrix(wb, photos, photos_with_cmt):
    ws = wb.create_sheet("🎨 월별 테마 매트릭스")
    ws.sheet_view.showGridLines = False
    _title_band(ws, "🎨 월별 테마 참여 예상 매트릭스", 15, height=32, size=13)

    ws.merge_cells("A2:O2")
    ws["A2"] = "▣ : 해당 월에 댓글 받은 사진을 1장 이상 업로드 (테마 이벤트 참여 가능성)"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="666666")
    ws["A2"].alignment = _left()
    ws.row_dimensions[2].height = 22

    user_month: dict = defaultdict(lambda: defaultdict(int))
    for p in photos_with_cmt:
        user_month[p["author"]][p["posted_at"].month] += 1

    mon_user_count: dict = defaultdict(int)
    for author, mons_d in user_month.items():
        for m in mons_d:
            mon_user_count[m] += 1

    def participate_months(author):
        return len(user_month[author])
    sorted_authors = sorted(
        user_month.keys(),
        key=lambda a: (-participate_months(a), -sum(user_month[a].values())),
    )

    ws.append([])
    header_row = 4
    ws.cell(header_row, 1).value = "작성자"
    ws.cell(header_row, 2).value = "참여월수"
    for i, m in enumerate(range(1, 13)):
        ws.cell(header_row, 3+i).value = f"{m}월"
    ws.cell(header_row, 15).value = "합계(장)"
    _style_header_row(ws, header_row, 1, 15)
    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = "C5"

    for author in sorted_authors:
        r = ws.max_row + 1
        mons_d = user_month[author]
        pm = participate_months(author)
        total_photos = sum(mons_d.values())
        ws.append([author, pm] + [mons_d.get(m, 0) for m in range(1, 13)] + [total_photos])
        bg = C["GRAY_LIGHT"] if r % 2 == 0 else C["WHITE"]
        for c in range(1, 16):
            cell = ws.cell(r, c)
            cell.font = _body_font(bold=(c == 1))
            cell.alignment = _left() if c == 1 else _center()
            cell.border = _thin_border()
            cell.fill = _fill(bg)
            if 3 <= c <= 14 and isinstance(cell.value, int) and cell.value > 0:
                cell.fill = _fill(C["THEME"])
                cell.font = Font(name="Arial", size=10, bold=True, color=C["ACCENT_PRP"])
                cell.value = f"▣ {cell.value}"
        ws.row_dimensions[r].height = 20

    # 월별 합계 행
    r = ws.max_row + 1
    ws.cell(r, 1).value = "월별 참여자 수"
    ws.cell(r, 2).value = "—"
    for i, m in enumerate(range(1, 13)):
        ws.cell(r, 3+i).value = mon_user_count.get(m, 0)
    ws.cell(r, 15).value = len(photos_with_cmt)
    for c in range(1, 16):
        cell = ws.cell(r, c)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = _fill(C["ACCENT_PRP"])
        cell.alignment = _center()
        cell.border = _thin_border()
    ws.row_dimensions[r].height = 22

    _set_col_widths(ws, {"A":12, "B":9,
                         **{get_column_letter(i): 7 for i in range(3, 15)},
                         "O":9})
    return user_month, mon_user_count, sorted_authors


def _build_sheet_photo_stats(wb, photos):
    ws = wb.create_sheet("👤 사진 통계")
    ws.sheet_view.showGridLines = False
    _title_band(ws, "👤 사용자별 사진 업로드 통계", 8, height=32, size=13)

    photo_stats: dict = defaultdict(lambda: {"total":0, "with_cmt":0, "likes":0, "comments":0})
    for p in photos:
        s = photo_stats[p["author"]]
        s["total"] += 1
        if p["has_comment"]:
            s["with_cmt"] += 1
        s["likes"]    += p["likes"]
        s["comments"] += p["comments"]
    sorted_photo_users = sorted(photo_stats.items(), key=lambda x: -x[1]["total"])

    hdrs = ["순위","작성자","사진수","테마예상","테마비율","좋아요합","댓글합","장당좋아요"]
    ws.append(hdrs)
    _style_header_row(ws, 2, 1, len(hdrs))
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A3"

    for rank, (author, s) in enumerate(sorted_photo_users, 1):
        r = ws.max_row + 1
        rate = s["with_cmt"]/s["total"]*100 if s["total"] else 0
        avg_likes = s["likes"]/s["total"] if s["total"] else 0
        ws.append([rank, author, s["total"], s["with_cmt"],
                   f"{rate:.1f}%", s["likes"], s["comments"], f"{avg_likes:.1f}"])
        bg = C["GRAY_LIGHT"] if rank % 2 == 0 else C["WHITE"]
        for c in range(1, len(hdrs)+1):
            cell = ws.cell(r, c)
            cell.font = _body_font(bold=(c == 2))
            cell.alignment = _left() if c == 2 else _center()
            cell.border = _thin_border()
            cell.fill = _fill(bg)
        if rate >= 30:
            ws.cell(r, 5).font = Font(name="Arial", size=10, bold=True, color=C["ACCENT_PRP"])
        ws.row_dimensions[r].height = 20

    n = len(sorted_photo_users)
    if n > 0:
        ws.conditional_formatting.add(
            f"C3:C{2+n}",
            DataBarRule(start_type="min", end_type="max", color="ED7D31", showValue=True),
        )
    _set_col_widths(ws, {"A":5,"B":12,"C":7,"D":8,"E":8,"F":8,"G":8,"H":10})

    return photo_stats, sorted_photo_users


def _build_sheet_insights(wb, posts, posts_A, posts_E, posts_active, posts_canceled,
                          photos, photos_with_cmt,
                          user_stats, sorted_users,
                          user_month, mon_user_count, sorted_authors,
                          photo_stats, sorted_photo_users,
                          period_label):
    ws = wb.create_sheet("💡 인사이트")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 22

    ws.merge_cells("B1:D1")
    ws["B1"] = f"💡 {GROUP_NAME} {period_label} 주요 인사이트"
    ws["B1"].font = _hdr_font(16, True)
    ws["B1"].fill = _fill(C["HDR_DARK"])
    ws["B1"].alignment = _center()
    ws.row_dimensions[1].height = 44

    def section(title, r):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        cell = ws.cell(r, 2)
        cell.value = title
        cell.font = _hdr_font(12, True)
        cell.fill = _fill(C["HDR_DARK"])
        cell.alignment = _left()
        ws.row_dimensions[r].height = 28
        return r + 1

    def row(label, value, note, r, hl=None):
        ws.row_dimensions[r].height = 22
        ws.cell(r, 2).value = label
        ws.cell(r, 3).value = value
        ws.cell(r, 4).value = note
        color = {
            "warn": C["ACCENT_YLW"], "good": "E2EFDA",
            "bad":  C["CANCEL"],    "theme": C["THEME"],
        }.get(hl, C["GRAY_LIGHT"] if r % 2 == 0 else C["WHITE"])
        for c in (2, 3, 4):
            cell = ws.cell(r, c)
            cell.fill = _fill(color)
            cell.font = _body_font(bold=(c == 2))
            cell.alignment = _left()
            cell.border = _thin_border()
        return r + 1

    def utotal(s): return s["A"] + s["A_취소"] + s["E"]

    r = 3
    r = section("① 활동량 & 규모", r)
    r = row("전체 게시글", f"{len(posts)}개", "공지+후기+가입인사", r)
    r = row("진행 출사", f"{len(posts_active)}건", "취소 제외", r, "good")
    if posts_A:
        r = row("취소 출사",
                f"{len(posts_canceled)}건 ({len(posts_canceled)/len(posts_A)*100:.1f}%)",
                "전체 공지 대비", r, "warn")
    r = row("사진 업로드", f"{len(photos)}장",
            f"월 평균 {len(photos)/12:.0f}장" if len(photos) else "—", r)
    if photos:
        r = row("테마 예상",
                f"{len(photos_with_cmt)}장 ({len(photos_with_cmt)/len(photos)*100:.1f}%)",
                "댓글 받은 사진", r, "theme")

    r += 1
    r = section("② 게시글 핵심 기여자 TOP 3", r)
    for rank, (author, s) in enumerate(sorted_users[:3], 1):
        tot = utotal(s)
        pct = tot/len(posts)*100 if posts else 0
        r = row(f"Top {rank}: {author}", f"{tot}건 ({pct:.1f}%)",
                f"공지 {s['A']} / 취소 {s['A_취소']} / 후기 {s['E']}",
                r, "good" if rank == 1 else None)

    r += 1
    r = section("③ 사진 업로드 TOP 5", r)
    for rank, (author, s) in enumerate(sorted_photo_users[:5], 1):
        rate = s["with_cmt"]/s["total"]*100 if s["total"] else 0
        r = row(f"Top {rank}: {author}", f"{s['total']}장",
                f"테마예상 {s['with_cmt']}장 ({rate:.0f}%) | 👍{s['likes']}", r)

    r += 1
    r = section("④ 테마 참여 패턴", r)
    r = row("테마 예상 참여자", f"{len(user_month)}명", "댓글 받은 사진 1장 이상", r, "theme")
    if sorted_authors:
        top = sorted_authors[0]
        r = row("최다 참여(개월수)", f"{top} - {len(user_month[top])}개월",
                "여러 달에 걸쳐 참여", r, "theme")
    if mon_user_count:
        hot = max(mon_user_count, key=mon_user_count.get)
        r = row("가장 활발한 달", f"{hot}월", f"{mon_user_count[hot]}명 참여", r, "good")

    r += 1
    r = section("⑤ 카테고리 트렌드 (출사 공지 기준)", r)
    all_cats = OUTING_CATS + NON_OUTING_CATS
    cat_counts = {c: sum(1 for p in posts_A if p["category"] == c) for c in all_cats}
    total_cat = sum(cat_counts.values())
    for cat in all_cats:
        cnt = cat_counts[cat]
        if not cnt:
            continue
        pct = cnt / total_cat * 100 if total_cat else 0
        kind = "출사" if cat in OUTING_CATS else "활동"
        r = row(f"[{cat}]", f"{cnt}건 ({pct:.1f}%)", kind, r)

    r += 1
    r = section("⑥ 취소율 분석 (공지 3건 이상)", r)
    cancel_users = [(a, s) for a, s in user_stats.items() if (s["A"]+s["A_취소"]) >= 3]
    cancel_users.sort(key=lambda x: -x[1]["A_취소"]/(x[1]["A"]+x[1]["A_취소"]))
    for author, s in cancel_users[:5]:
        n = s["A"] + s["A_취소"]
        pct = s["A_취소"]/n*100
        hl = "bad" if pct >= 50 else ("warn" if pct >= 25 else None)
        r = row(f"{author}", f"{pct:.1f}% ({s['A_취소']}/{n})", "", r, hl)


# ═══════════════════════════════════════════════════════════════
# 후기 본문 기반 참석 시트 (PR3)
# ═══════════════════════════════════════════════════════════════
# 모든 헬퍼는 attendees/actually_held/matched_review_id 키가 없을 때도(KeyError 없이)
# 빈 결과를 내야 한다 — save_excel을 annotate 없이 호출하는 코드경로 보존.

def _build_sheet_outing_attendees(wb, posts_A):
    ws = wb.create_sheet("🎯 출사별 참석자")
    ws.sheet_view.showGridLines = False
    _title_band(ws, "🎯 출사별 참석자 — 후기 본문 기반", 7, height=32, size=13)

    hdrs = ["출사날짜", "카테고리", "공지자", "제목", "참석자수", "매칭", "참석자 명단"]
    ws.append(hdrs)
    _style_header_row(ws, 2, 1, len(hdrs), bg=C["HDR_MID"])
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    for p in sorted(posts_A, key=lambda x: x.get("outing_date") or "0000", reverse=True):
        att = p.get("attendees", []) or []
        matched = "✓" if p.get("matched_review_id") else "—"
        r = ws.max_row + 1
        ws.append([
            p.get("outing_date") or "-",
            p.get("category") or "-",
            p.get("author", ""),
            p.get("title", ""),
            len(att),
            matched,
            ", ".join(att) if att else "—",
        ])
        if p.get("is_canceled"):
            bg = C["CANCEL"]
        elif p.get("actually_held"):
            bg = C["OUTING"]
        else:
            bg = C["GRAY_LIGHT"]
        for c in range(1, len(hdrs) + 1):
            cell = ws.cell(r, c)
            cell.font = _body_font()
            cell.fill = _fill(bg)
            cell.border = _thin_border()
            cell.alignment = _left() if c in (4, 7) else _center()
        ws.row_dimensions[r].height = 18

    _set_col_widths(ws, {"A": 12, "B": 11, "C": 11, "D": 36, "E": 8, "F": 6, "G": 50})


def _build_sheet_member_attendance(wb, posts_A):
    ws = wb.create_sheet("👥 멤버별 참석")
    ws.sheet_view.showGridLines = False
    _title_band(ws, "👥 멤버별 참석 통계 (정규화 카테고리 동적 컬럼)", 11, height=32, size=13)

    cats = list(OUTING_CATS) + list(NON_OUTING_CATS)
    hdrs = ["순위", "멤버", "총 참석"] + cats + ["첫 등장", "최근 등장"]
    ws.append(hdrs)
    _style_header_row(ws, 2, 1, len(hdrs), bg=C["HDR_MID"])
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "C3"

    from collections import Counter, defaultdict
    attend: Counter = Counter()
    cat_pref: dict = defaultdict(Counter)
    first: dict = {}
    last: dict = {}
    held = sorted(
        (p for p in posts_A if p.get("actually_held") and p.get("outing_date")),
        key=lambda x: x["outing_date"],
    )
    for n in held:
        cat = n.get("category")
        for name in n.get("attendees", []):
            attend[name] += 1
            if cat:
                cat_pref[name][cat] += 1
            first.setdefault(name, n["outing_date"])
            last[name] = n["outing_date"]

    ranked = attend.most_common()
    max_total = ranked[0][1] if ranked else 0
    for rank, (name, total) in enumerate(ranked, 1):
        r = ws.max_row + 1
        cat_counts = [cat_pref[name].get(c, 0) for c in cats]
        ws.append([rank, name, total] + cat_counts + [first.get(name, "-"), last.get(name, "-")])
        bg = C["GRAY_LIGHT"] if rank % 2 == 0 else C["WHITE"]
        for c in range(1, len(hdrs) + 1):
            cell = ws.cell(r, c)
            cell.font = _body_font(bold=(c == 2))
            cell.alignment = _left() if c == 2 else _center()
            cell.border = _thin_border()
            cell.fill = _fill(bg)
        ws.row_dimensions[r].height = 20

    if ranked:
        last_row = 2 + len(ranked)
        ws.conditional_formatting.add(
            f"C3:C{last_row}",
            DataBarRule(start_type="min", end_type="max", color="2E75B6", showValue=True),
        )

    widths = {"A": 5, "B": 11, "C": 8}
    for i, _c in enumerate(cats):
        widths[get_column_letter(4 + i)] = 11
    widths[get_column_letter(4 + len(cats))] = 12      # 첫 등장
    widths[get_column_letter(5 + len(cats))] = 12      # 최근 등장
    _set_col_widths(ws, widths)


def _build_sheet_monthly_matrix(wb, posts_A):
    ws = wb.create_sheet("📅 월별 참석 매트릭스")
    ws.sheet_view.showGridLines = False
    _title_band(ws, "📅 월별 참석 매트릭스 — 멤버 × 월", 15, height=32, size=13)

    ws.merge_cells("A2:O2")
    ws["A2"] = "▣ : 해당 월 출사 후기에서 참석자로 매칭됨"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="666666")
    ws["A2"].alignment = _left()
    ws.row_dimensions[2].height = 22

    from collections import defaultdict
    mm: dict = defaultdict(lambda: defaultdict(int))
    total_per_month: dict = defaultdict(int)
    for n in posts_A:
        if not n.get("actually_held") or not n.get("outing_date"):
            continue
        m = date.fromisoformat(n["outing_date"]).month
        for name in n.get("attendees", []):
            mm[name][m] += 1
            total_per_month[m] += 1

    sorted_members = sorted(
        mm.keys(),
        key=lambda a: (-len([m for m in mm[a] if mm[a][m] > 0]),
                       -sum(mm[a].values())),
    )

    header_row = 4
    ws.cell(header_row, 1).value = "멤버"
    ws.cell(header_row, 2).value = "참석월수"
    for i, m in enumerate(range(1, 13)):
        ws.cell(header_row, 3 + i).value = f"{m}월"
    ws.cell(header_row, 15).value = "합계"
    _style_header_row(ws, header_row, 1, 15)
    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = "C5"

    for member in sorted_members:
        r = ws.max_row + 1
        months = mm[member]
        pm = sum(1 for m in months if months[m] > 0)
        total = sum(months.values())
        ws.append([member, pm] + [months.get(m, 0) for m in range(1, 13)] + [total])
        bg = C["GRAY_LIGHT"] if r % 2 == 0 else C["WHITE"]
        for c in range(1, 16):
            cell = ws.cell(r, c)
            cell.font = _body_font(bold=(c == 1))
            cell.alignment = _left() if c == 1 else _center()
            cell.border = _thin_border()
            cell.fill = _fill(bg)
            if 3 <= c <= 14 and isinstance(cell.value, int) and cell.value > 0:
                cell.fill = _fill(C["OUTING"])
                cell.font = Font(name="Arial", size=10, bold=True, color=C["ACCENT_GRN"])
                cell.value = f"▣ {cell.value}"
        ws.row_dimensions[r].height = 20

    # 월별 합계 행
    if sorted_members:
        r = ws.max_row + 1
        ws.cell(r, 1).value = "월별 참석 연인원"
        ws.cell(r, 2).value = "—"
        for i, m in enumerate(range(1, 13)):
            ws.cell(r, 3 + i).value = total_per_month.get(m, 0)
        ws.cell(r, 15).value = sum(total_per_month.values())
        for c in range(1, 16):
            cell = ws.cell(r, c)
            cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            cell.fill = _fill(C["ACCENT_GRN"])
            cell.alignment = _center()
            cell.border = _thin_border()
        ws.row_dimensions[r].height = 22

    widths = {"A": 12, "B": 9}
    for i in range(12):
        widths[get_column_letter(3 + i)] = 8
    widths["O"] = 9
    _set_col_widths(ws, widths)
